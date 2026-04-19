"""
Build synthetic RVTools-style .xlsx workbooks from a template (e.g. customer.xlsx).
Strategy: subset VMs by size, filter dependent rows on all sheets, then apply
consistent placeholder names for VMs, hosts, clusters and datacenters.
XXL: keep all template VMs and append synthetic clones (40% extra) by copying
per-VM rows and renaming the seed VM string to a unique clone token.
"""

from __future__ import annotations

import math
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

# Target VM count per size (template vInfo row count is the upper bound except XXL)
SIZE_VM_COUNTS: dict[str, int | None] = {
    "s": 120,
    "m": 400,
    "l": 1000,
    "xl": None,  # all VMs in the template
}

VALID_SIZES = frozenset({*SIZE_VM_COUNTS, "xxl"})

# XXL: number of additional synthetic VMs = ceil(original_vm_count * 0.4)
XXL_EXTRA_VM_FRACTION = 0.4

_DS_BRACKET = re.compile(r"\[([^\]]+)\]")


def _norm_size(size: str) -> str:
    s = (size or "").strip().lower()
    if s not in VALID_SIZES:
        raise ValueError("size must be one of: s, m, l, xl, xxl")
    return s


def _substitute_seed_with_clone(row: tuple[Any, ...], seed: Any, clone_token: str) -> tuple[Any, ...]:
    seed_s = str(seed)
    out: list[Any] = []
    for c in row:
        if isinstance(c, str):
            out.append(c.replace(seed_s, clone_token))
        elif c is not None and str(c) == seed_s:
            out.append(clone_token)
        else:
            out.append(c)
    return tuple(out)


def _append_cloned_vm_rows(
    header: tuple[Any, ...],
    data: list[tuple[Any, ...]],
    vm_selected: list[Any],
    clone_tokens: list[str],
) -> list[tuple[Any, ...]]:
    if not header or "VM" not in header or not clone_tokens:
        return data
    vi = list(header).index("VM")
    by_vm: dict[Any, list[tuple[Any, ...]]] = defaultdict(list)
    for r in data:
        if len(r) > vi and r[vi] is not None:
            by_vm[r[vi]].append(r)
    extra: list[tuple[Any, ...]] = []
    n_orig = len(vm_selected)
    for j, tok in enumerate(clone_tokens):
        seed = vm_selected[j % n_orig]
        for r in by_vm.get(seed, []):
            extra.append(_substitute_seed_with_clone(r, seed, tok))
    return list(data) + extra


def _extract_datastore_names_from_path(path: str | None) -> set[str]:
    if not path or not isinstance(path, str):
        return set()
    m = _DS_BRACKET.search(path)
    if not m:
        return set()
    return {m.group(1)}


def _collect_rows(ws) -> tuple[tuple[Any, ...], list[tuple[Any, ...]]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return (), []
    return rows[0], rows[1:]


def _filter_vm_rows(
    header: tuple[Any, ...], data: list[tuple[Any, ...]], vm_set: set[Any]
) -> list[tuple[Any, ...]]:
    if not header or "VM" not in header:
        return data
    idx = list(header).index("VM")
    return [r for r in data if len(r) > idx and r[idx] in vm_set]


def _filter_host_rows(
    header: tuple[Any, ...], data: list[tuple[Any, ...]], host_set: set[Any]
) -> list[tuple[Any, ...]]:
    if not header or "Host" not in header:
        return data
    idx = list(header).index("Host")
    return [r for r in data if len(r) > idx and r[idx] in host_set]


def _filter_cluster_rows(
    header: tuple[Any, ...], data: list[tuple[Any, ...]], cluster_set: set[Any]
) -> list[tuple[Any, ...]]:
    if not header or "Name" not in header:
        return data
    idx = list(header).index("Name")
    return [r for r in data if len(r) > idx and r[idx] in cluster_set]


def _filter_datastore_rows(
    header: tuple[Any, ...], data: list[tuple[Any, ...]], ds_set: set[Any]
) -> list[tuple[Any, ...]]:
    if not header or "Name" not in header:
        return data
    idx = list(header).index("Name")
    return [r for r in data if len(r) > idx and r[idx] in ds_set]


def _filter_multipath_rows(
    header: tuple[Any, ...],
    data: list[tuple[Any, ...]],
    host_set: set[Any],
    ds_set: set[Any],
) -> list[tuple[Any, ...]]:
    if not header:
        return data
    h = list(header)
    if "Host" not in h or "Datastore" not in h:
        return data
    hi = h.index("Host")
    di = h.index("Datastore")
    out = []
    for r in data:
        if len(r) <= max(hi, di):
            continue
        if r[hi] in host_set and r[di] in ds_set:
            out.append(r)
    return out


def _slice_rows(data: list[tuple[Any, ...]], max_rows: int) -> list[tuple[Any, ...]]:
    if max_rows <= 0:
        return []
    return data[:max_rows]


def _build_replacement_maps(
    vm_selected: list[Any],
    host_set: set[Any],
    cluster_set: set[Any],
    dc_set: set[Any],
) -> tuple[dict[str, str], list[tuple[str, str]]]:
    vm_map = {str(v): f"SYN-VM-{i:05d}" for i, v in enumerate(vm_selected, start=1)}
    host_sorted = sorted(str(h) for h in host_set if h is not None)
    cluster_sorted = sorted(str(c) for c in cluster_set if c is not None)
    dc_sorted = sorted(str(d) for d in dc_set if d is not None)

    host_map = {h: f"syn-esxi-{i:03d}.local" for i, h in enumerate(host_sorted, start=1)}
    cluster_map = {c: f"SYN-CLUSTER-{i:03d}" for i, c in enumerate(cluster_sorted, start=1)}
    dc_map = {d: f"SYN-DC-{i:02d}" for i, d in enumerate(dc_sorted, start=1)}

    ordered: list[tuple[str, str]] = []
    ordered.extend((k, v) for k, v in vm_map.items())
    ordered.extend((k, v) for k, v in host_map.items())
    ordered.extend((k, v) for k, v in cluster_map.items())
    ordered.extend((k, v) for k, v in dc_map.items())
    ordered.sort(key=lambda kv: len(kv[0]), reverse=True)
    return vm_map, ordered


def _anonymize_cell(val: Any, replacements: list[tuple[str, str]]) -> Any:
    if not isinstance(val, str) or not val:
        return val
    s = val
    for old, new in replacements:
        if old in s:
            s = s.replace(old, new)
    return s


def _rewrite_sheet(ws, header: tuple[Any, ...], data: list[tuple[Any, ...]], replacements: list[tuple[str, str]]):
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    if header:
        ws.append([_anonymize_cell(c, replacements) for c in header])
    for row in data:
        ws.append([_anonymize_cell(c, replacements) for c in row])


def generate_rvtools_xlsx(
    template_path: str | Path,
    output_path: str | Path,
    size: str,
) -> dict[str, Any]:
    """Read template workbook, filter / clone by size, write a new .xlsx file."""
    template_path = Path(template_path)
    output_path = Path(output_path)
    key = _norm_size(size)

    wb_r = load_workbook(template_path, read_only=True, data_only=True)
    if "vInfo" not in wb_r.sheetnames:
        wb_r.close()
        raise ValueError("Template has no vInfo sheet")

    vinfo = wb_r["vInfo"]
    vinfo_header, vinfo_data = _collect_rows(vinfo)
    if not vinfo_header or "VM" not in vinfo_header:
        wb_r.close()
        raise ValueError("vInfo sheet has no VM column")

    vm_idx = list(vinfo_header).index("VM")
    all_vms_ordered = [r[vm_idx] for r in vinfo_data if len(r) > vm_idx and r[vm_idx] is not None]

    if key in ("xl", "xxl"):
        target_n = len(all_vms_ordered)
    else:
        target_n = int(SIZE_VM_COUNTS[key])  # s, m, l
    target_n = max(1, min(target_n, len(all_vms_ordered)))
    vm_selected = all_vms_ordered[:target_n]
    vm_set = set(vm_selected)

    clone_tokens: list[str] = []
    if key == "xxl":
        extra = max(1, int(math.ceil(len(vm_selected) * XXL_EXTRA_VM_FRACTION)))
        clone_tokens = [f"__SYN_CLONE_{i:05d}__" for i in range(extra)]

    h_cpu, vcpu_data = _collect_rows(wb_r["vCPU"])
    if "VM" not in h_cpu or "Host" not in h_cpu or "Cluster" not in h_cpu or "Datacenter" not in h_cpu:
        wb_r.close()
        raise ValueError("vCPU sheet is missing expected columns")

    vi = list(h_cpu).index("VM")
    hi = list(h_cpu).index("Host")
    ci = list(h_cpu).index("Cluster")
    di = list(h_cpu).index("Datacenter")

    host_set: set[Any] = set()
    cluster_set: set[Any] = set()
    dc_set: set[Any] = set()
    for r in vcpu_data:
        if len(r) <= max(vi, hi, ci, di):
            continue
        if r[vi] not in vm_set:
            continue
        host_set.add(r[hi])
        cluster_set.add(r[ci])
        dc_set.add(r[di])

    h_disk, vdisk_data = _collect_rows(wb_r["vDisk"])
    if "VM" not in h_disk:
        wb_r.close()
        raise ValueError("vDisk sheet has no VM column")
    vdi = list(h_disk).index("VM")
    path_candidates = [c for c in ("Path", "Disk Path") if c in h_disk]
    pi = list(h_disk).index(path_candidates[0]) if path_candidates else None

    ds_set: set[str] = set()
    for r in vdisk_data:
        if len(r) <= vdi or r[vdi] not in vm_set:
            continue
        if pi is not None and len(r) > pi:
            ds_set |= _extract_datastore_names_from_path(r[pi])

    if not ds_set:
        # Fallback: all template datastores if paths are not parseable
        ds = wb_r["vDatastore"]
        h_ds, ds_data = _collect_rows(ds)
        if h_ds and "Name" in h_ds:
            ni = list(h_ds).index("Name")
            ds_set = {r[ni] for r in ds_data if len(r) > ni and r[ni] is not None}

    wb_r.close()

    vm_selected_for_map = list(vm_selected) + clone_tokens
    _, replacements = _build_replacement_maps(vm_selected_for_map, host_set, cluster_set, dc_set)

    wb_w = load_workbook(template_path, read_only=False, data_only=True)
    new_wb = Workbook()
    default = new_wb.active
    new_wb.remove(default)

    for name in wb_w.sheetnames:
        src = wb_w[name]
        header, data = _collect_rows(src)
        if not header:
            dst = new_wb.create_sheet(title=name)
            continue

        if name == "vCluster":
            data = _filter_cluster_rows(header, data, cluster_set)
        elif name == "vDatastore":
            data = _filter_datastore_rows(header, data, ds_set)
        elif name == "vMultiPath":
            data = _filter_multipath_rows(header, data, host_set, ds_set)
        elif name in {"vHost", "vHBA", "vNIC", "vSwitch", "vPort", "vSC_VMK"}:
            data = _filter_host_rows(header, data, host_set)
        elif name in {"vSource", "vRP"}:
            cap = max(40, min(len(data), len(vm_set) * 3))
            data = _slice_rows(data, cap)
        elif name in {"dvSwitch", "dvPort"}:
            data = _slice_rows(data, min(len(data), 400))
        elif "VM" in header:
            data = _filter_vm_rows(header, data, vm_set)
            if clone_tokens:
                data = _append_cloned_vm_rows(header, data, vm_selected, clone_tokens)
        else:
            data = _slice_rows(data, min(len(data), 200))

        dst = new_wb.create_sheet(title=name)
        _rewrite_sheet(dst, header, data, replacements)

    wb_w.close()
    new_wb.save(output_path)

    total_vms = len(vm_selected) + len(clone_tokens)
    out: dict[str, Any] = {
        "size": key,
        "vm_count": total_vms,
        "hosts": len(host_set),
        "clusters": len(cluster_set),
        "datastores_named": len(ds_set),
        "output": str(output_path),
    }
    if clone_tokens:
        out["original_vm_count"] = len(vm_selected)
        out["synthetic_clones"] = len(clone_tokens)
    return out
